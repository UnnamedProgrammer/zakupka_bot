from io import BytesIO
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import select, inspect
from sqlalchemy.orm import selectinload

from app.db.models import Attachment, Request, RequestItem
from app.services.files import save_bytes_file


class TemplateParseError(Exception):
    pass


class ReportParseError(Exception):
    pass


REQUEST_TEMPLATE_PATH = Path(__file__).resolve().parents[1] / "templates" / "request_template.xlsx"


REQUESTS_REPORT_HEADERS = [
    "ID",
    "Дата создания",
    "Дата обновления",
    "Инициатор",
    "Подразделение",
    "ЦФО",
    "МОЛ",
    "Статус",
    "Исполнитель",
    "Поставщик",
    "Срок поставки",
    "Дата согласования",
    "Дата выполнения",
    "Дата получения",
    "Способ описания",
    "Товары",
    "Комментарии",
]


def _get_loaded(obj, attr: str):
    state = inspect(obj)
    if attr in state.unloaded:
        return None
    return getattr(obj, attr)


def _normalize_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    return " ".join(text.split())


def _normalize_header(value) -> str:
    return _normalize_cell(value).casefold()


def _clean_optional_cell(value) -> str | None:
    text = _normalize_cell(value)
    if not text:
        return None
    lowered = text.casefold()
    if lowered in {"-", "—", "нет", "пропустить", "skip"}:
        return None
    return text


def _ensure_merged(ws, cell_range: str, force: bool = False) -> None:
    try:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    except ValueError:
        return
    overlaps: list[str] = []
    for merged in ws.merged_cells.ranges:
        if merged.coord == cell_range:
            return
        if (
            min_row <= merged.max_row
            and max_row >= merged.min_row
            and min_col <= merged.max_col
            and max_col >= merged.min_col
        ):
            if (
                min_row >= merged.min_row
                and max_row <= merged.max_row
                and min_col >= merged.min_col
                and max_col <= merged.max_col
            ):
                return
            if not force:
                return
            if (
                min_row <= merged.min_row
                and max_row >= merged.max_row
                and min_col <= merged.min_col
                and max_col >= merged.max_col
            ):
                overlaps.append(merged.coord)
            else:
                return
    for coord in overlaps:
        ws.unmerge_cells(coord)
    ws.merge_cells(cell_range)


def parse_request_template(path: str) -> dict:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    initiator_name = _normalize_cell(ws["E2"].value)
    if not initiator_name:
        raise TemplateParseError("В ячейке E2 не указан инициатор заявки.")

    template_department = _normalize_cell(ws["E4"].value)
    if not template_department:
        raise TemplateParseError("В ячейке E4 не указано подразделение.")

    template_cfo = _normalize_cell(ws["E5"].value)
    if not template_cfo:
        raise TemplateParseError("В ячейке E5 не указано ЦФО.")

    rows_found = 0
    mol_full_name: str | None = None
    contract_max_price: str | None = None
    bdds_article_category: str | None = None
    items: list[dict] = []

    for row_idx in range(7, ws.max_row + 1):
        raw_values = [ws.cell(row=row_idx, column=col).value for col in range(1, 9)]
        values = [_normalize_cell(value) for value in raw_values]
        if not any(values):
            if rows_found:
                break
            continue
        rows_found += 1

        row_mol, name, specs, qty, unit, note, max_price, bdds_value = values
        cleaned_specs = _clean_optional_cell(raw_values[2])
        cleaned_note = _clean_optional_cell(raw_values[5])
        if not name:
            raise TemplateParseError(f"Строка {row_idx}: не заполнено \"Наименование\".")
        if not qty:
            raise TemplateParseError(f"Строка {row_idx}: не заполнено \"Количество\".")
        if not unit:
            raise TemplateParseError(f"Строка {row_idx}: не заполнено \"Ед. измерения\".")
        if not row_mol:
            raise TemplateParseError(f"Строка {row_idx}: не заполнено \"МОЛ\".")

        if mol_full_name is None:
            mol_full_name = row_mol
        elif mol_full_name.casefold() != row_mol.casefold():
            raise TemplateParseError(
                f"Строка {row_idx}: указан другой МОЛ. Ожидается: {mol_full_name}."
            )

        row_price = _clean_optional_cell(raw_values[6])
        if row_price:
            if contract_max_price is None:
                contract_max_price = row_price
            elif contract_max_price.casefold() != row_price.casefold():
                raise TemplateParseError(
                    f"Строка {row_idx}: указана другая макс. цена договора. "
                    f"Ожидается: {contract_max_price}."
                )

        row_bdds = _clean_optional_cell(raw_values[7])
        if row_bdds:
            if bdds_article_category is None:
                bdds_article_category = row_bdds
            elif bdds_article_category.casefold() != row_bdds.casefold():
                raise TemplateParseError(
                    f"Строка {row_idx}: указано другое значение БДДС. "
                    f"Ожидается: {bdds_article_category}."
                )

        items.append(
            {
                "name": name,
                "specs": cleaned_specs,
                "brand": None,
                "qty": qty,
                "unit": unit,
                "link": None,
                "note": cleaned_note,
                "omts_responsible": None,
                "category": None,
                "dds_article": None,
                "max_price": None,
                "attachments": [],
            }
        )

    if not items:
        raise TemplateParseError("В файле не найдены строки с товарами (начиная с 7-й строки).")

    return {
        "initiator_name": initiator_name,
        "groups": [
            {
                "department_name": template_department,
                "cfo_name": template_cfo,
                "mol_full_name": mol_full_name,
                "contract_max_price": contract_max_price,
                "bdds_article_category": bdds_article_category,
                "items": items,
            }
        ],
    }


def build_request_template_xlsx(
    request: Request,
    items: Iterable[RequestItem],
    template_path: str | Path = REQUEST_TEMPLATE_PATH,
) -> bytes:
    path = Path(template_path)
    if not path.exists():
        raise FileNotFoundError(f"Template not found: {path}")
    wb = load_workbook(path)
    ws = wb.active

    initiator = _get_loaded(request, "initiator")
    department = _get_loaded(request, "department")
    cfo = _get_loaded(request, "cfo")

    ws["E2"] = initiator.full_name if initiator else ""
    ws["E4"] = department.name if department else ""
    ws["E5"] = cfo.name if cfo else ""
    ws["E2"].alignment = Alignment(vertical="center", wrap_text=True)
    ws["E4"].alignment = Alignment(vertical="center", wrap_text=True)
    ws["E5"].alignment = Alignment(vertical="center", wrap_text=True)

    mol_name = request.mol_full_name or ""
    contract_max_price = request.contract_max_price or ""
    bdds_article_category = request.bdds_article_category or ""

    items_list = list(items or [])
    if not items_list and (
        request.item_name
        or request.item_specs
        or request.item_qty
        or request.item_unit
        or request.item_note
    ):
        items_list = [
            RequestItem(
                name=request.item_name,
                specs=request.item_specs,
                qty=request.item_qty,
                unit=request.item_unit,
                note=request.item_note,
            )
        ]

    row_idx = 7
    for item in items_list:
        ws.cell(row=row_idx, column=1, value=mol_name)
        ws.cell(row=row_idx, column=2, value=item.name)
        ws.cell(row=row_idx, column=3, value=item.specs)
        ws.cell(row=row_idx, column=4, value=item.qty)
        ws.cell(row=row_idx, column=5, value=item.unit)
        ws.cell(row=row_idx, column=6, value=item.note)
        ws.cell(row=row_idx, column=7, value=contract_max_price)
        ws.cell(row=row_idx, column=8, value=bdds_article_category)
        row_idx += 1

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_request_template_prefilled_xlsx(
    department_name: str | None,
    cfo_name: str | None,
    initiator_name: str | None = None,
    template_path: str | Path = REQUEST_TEMPLATE_PATH,
) -> bytes:
    path = Path(template_path)
    if not path.exists():
        raise FileNotFoundError(f"Template not found: {path}")
    wb = load_workbook(path)
    ws = wb.active

    if initiator_name:
        ws["E2"] = initiator_name
    if department_name:
        ws["E4"] = department_name
        ws["E4"].alignment = Alignment(vertical="center", wrap_text=True)
    if cfo_name:
        ws["E5"] = cfo_name
        ws["E5"].alignment = Alignment(vertical="center", wrap_text=True)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def parse_requests_report_xlsx(path: str) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    header_map: dict[str, int] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        key = _normalize_header(cell.value)
        if not key:
            continue
        if key in header_map:
            raise ReportParseError(f"Дублирующийся заголовок: {cell.value}")
        header_map[key] = col_idx

    missing = [
        header
        for header in REQUESTS_REPORT_HEADERS
        if _normalize_header(header) not in header_map
    ]
    if missing:
        raise ReportParseError(
            "Не найдены заголовки: " + ", ".join(missing)
        )

    rows: list[dict] = []
    for row_idx in range(2, ws.max_row + 1):
        values: dict[str, object | None] = {}
        is_empty = True
        for header in REQUESTS_REPORT_HEADERS:
            col_idx = header_map[_normalize_header(header)]
            value = ws.cell(row=row_idx, column=col_idx).value
            if _normalize_cell(value):
                is_empty = False
            values[header] = value
        if is_empty:
            continue
        rows.append({"row": row_idx, "values": values})

    if not rows:
        raise ReportParseError("В файле нет строк для обработки.")

    return rows


def _build_items_text(req: Request) -> str:
    items = _get_loaded(req, "items") or []
    if items:
        item_blocks = []
        for idx, item in enumerate(items, start=1):
            omts = _get_loaded(item, "omts_responsible")
            category = _get_loaded(item, "category")
            dds = _get_loaded(item, "dds_article")
            item_blocks.append(
                "\n".join(
                    [
                        f"{idx}. Наименование: {item.name or ''}",
                        f"   Характеристики: {item.specs or ''}",
                        f"   Марка/аналог: {item.brand or ''}",
                        f"   Количество: {item.qty or ''}",
                        f"   Ед.: {item.unit or ''}",
                        f"   Ссылка: {item.link or '-'}",
                        f"   Примечание: {item.note or ''}",
                        f"   Ответственный ОМТС: {omts.name if omts else '-'}",
                        f"   Категория: {category.name if category else '-'}",
                        f"   Статья ДДС: {dds.name if dds else '-'}",
                        f"   Макс. цена: {item.max_price or '-'}",
                    ]
                ).strip()
            )
        return "\n\n".join(item_blocks).strip()
    if req.item_name or req.item_specs or req.item_brand or req.item_qty or req.item_unit:
        return "\n".join(
            [
                f"1. Наименование: {req.item_name or ''}",
                f"   Характеристики: {req.item_specs or ''}",
                f"   Марка/аналог: {req.item_brand or ''}",
                f"   Количество: {req.item_qty or ''}",
                f"   Ед.: {req.item_unit or ''}",
                f"   Ссылка: {req.item_link or '-'}",
                f"   Примечание: {req.item_note or ''}",
            ]
        ).strip()
    return ""


def _build_comments_text(req: Request) -> str:
    comments = _get_loaded(req, "comments") or []
    lines = [comment.text for comment in comments if comment.text]
    return "\n".join(lines).strip()


def _apply_table_formatting(ws, header_len: int) -> None:
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    alt_fill = PatternFill("solid", fgColor="F7F7F7")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, header_len + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = bold
        cell.alignment = center
        cell.fill = header_fill
        cell.border = border

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=header_len), start=2):
        for cell in row:
            cell.alignment = wrap
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    last_col = get_column_letter(header_len)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"


def build_archive_xlsx(requests: Iterable[Request]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Архив заявок"

    header = [
        "ID",
        "Дата создания",
        "Инициатор",
        "Подразделение",
        "ЦФО",
        "МОЛ",
        "Статус",
        "Поставщик",
        "Товары",
    ]
    ws.append(header)

    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    alt_fill = PatternFill("solid", fgColor="F7F7F7")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = bold
        cell.alignment = center
        cell.fill = header_fill
        cell.border = border

    for req in requests:
        items = list(getattr(req, "items", []) or [])
        if not items:
            items = [
                RequestItem(
                    name=req.item_name,
                    specs=req.item_specs,
                    brand=req.item_brand,
                    qty=req.item_qty,
                    unit=req.item_unit,
                    link=req.item_link,
                    note=req.item_note,
                )
            ]
        item_blocks = []
        for idx, item in enumerate(items, start=1):
            item_blocks.append(
                "\n".join(
                    [
                        f"{idx}. Наименование: {item.name or ''}",
                        f"   Характеристики: {item.specs or ''}",
                        f"   Марка/аналог: {item.brand or ''}",
                        f"   Количество: {item.qty or ''}",
                        f"   Ед.: {item.unit or ''}",
                        f"   Ссылка: {item.link or '-'}",
                        f"   Примечание: {item.note or ''}",
                    ]
                ).strip()
            )
        items_text = "\n\n".join(item_blocks).strip()
        ws.append(
            [
                req.id,
                req.created_at.strftime("%Y-%m-%d %H:%M") if req.created_at else "",
                req.initiator.full_name if req.initiator else "",
                req.department.name if req.department else "",
                req.cfo.name if req.cfo else "",
                req.mol_full_name or "",
                req.status.name if req.status else "",
                req.supplier_name or "",
                items_text,
            ]
        )

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=len(header)), start=2):
        for cell in row:
            cell.alignment = wrap
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 40
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{ws.max_row}"

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _build_requests_report_xlsx(requests: Iterable[Request], title: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title

    ws.append(REQUESTS_REPORT_HEADERS)

    for req in requests:
        created_at = req.created_at.strftime("%Y-%m-%d %H:%M") if req.created_at else ""
        updated_at = req.updated_at.strftime("%Y-%m-%d %H:%M") if req.updated_at else ""
        approved_at = req.approved_at.strftime("%Y-%m-%d %H:%M") if req.approved_at else ""
        done_at = req.done_at.strftime("%Y-%m-%d %H:%M") if req.done_at else ""
        received_at = req.received_at.strftime("%Y-%m-%d %H:%M") if req.received_at else ""
        expected_delivery = (
            req.expected_delivery_at.strftime("%Y-%m-%d") if req.expected_delivery_at else ""
        )
        ws.append(
            [
                req.id,
                created_at,
                updated_at,
                req.initiator.full_name if req.initiator else "",
                req.department.name if req.department else "",
                req.cfo.name if req.cfo else "",
                req.mol_full_name or "",
                req.status.name if req.status else "",
                req.executor.full_name if req.executor else "",
                req.supplier_name or "",
                expected_delivery,
                approved_at,
                done_at,
                received_at,
                req.description_method or "",
                _build_items_text(req),
                _build_comments_text(req),
            ]
        )

    _apply_table_formatting(ws, len(REQUESTS_REPORT_HEADERS))

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 22
    ws.column_dimensions["J"].width = 20
    ws.column_dimensions["K"].width = 14
    ws.column_dimensions["L"].width = 18
    ws.column_dimensions["M"].width = 18
    ws.column_dimensions["N"].width = 18
    ws.column_dimensions["O"].width = 16
    ws.column_dimensions["P"].width = 55
    ws.column_dimensions["Q"].width = 40

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_archive_requests_xlsx(requests: Iterable[Request]) -> bytes:
    return _build_requests_report_xlsx(requests, "Архив заявок")


def build_daily_requests_xlsx(requests: Iterable[Request]) -> bytes:
    return _build_requests_report_xlsx(requests, "Ежедневные заявки")


def build_employee_stats_xlsx(requests: Iterable[Request]) -> bytes:
    return _build_requests_report_xlsx(requests, "Статистика сотрудников")


def build_request_xlsx(
    request: Request, items: Iterable[RequestItem], attachments: Iterable[Attachment]
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"Заявка {request.id}"

    header_fill = PatternFill("solid", fgColor="D9E1F2")
    section_fill = PatternFill("solid", fgColor="F2F2F2")
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    created_at = request.created_at.strftime("%Y-%m-%d %H:%M") if request.created_at else ""
    ws.append(
        [
            "Заявка №",
            "Дата создания",
            "Инициатор",
            "Подразделение",
            "ЦФО",
            "МОЛ",
            "Статус",
        ]
    )
    ws.append(
        [
            request.id,
            created_at,
            request.initiator.full_name if request.initiator else "",
            request.department.name if request.department else "",
            request.cfo.name if request.cfo else "",
            request.mol_full_name or "",
            request.status.name if request.status else "",
        ]
    )
    ws.append([])

    ws.append(
        [
            "№",
            "Наименование",
            "Характеристики",
            "Марка/аналог",
            "Количество",
            "Ед.",
            "Ссылка",
            "Примечание",
        ]
    )
    header_row = ws.max_row
    for col in range(1, 9):
        cell = ws.cell(row=header_row, column=col)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    item_list = list(items)
    if not item_list:
        ws.append(
            [
                1,
                request.item_name or "",
                request.item_specs or "",
                request.item_brand or "",
                request.item_qty or "",
                request.item_unit or "",
                request.item_link or "-",
                request.item_note or "",
            ]
        )
    else:
        for idx, item in enumerate(item_list, start=1):
            ws.append(
                [
                    idx,
                    item.name or "",
                    item.specs or "",
                    item.brand or "",
                    item.qty or "",
                    item.unit or "",
                    item.link or "-",
                    item.note or "",
                ]
            )
    for row in ws.iter_rows(min_row=header_row + 1, max_col=8):
        for cell in row:
            cell.alignment = wrap
            cell.border = border

    for col in range(1, 8):
        cell = ws.cell(row=1, column=col)
        cell.font = bold
        cell.fill = section_fill
        cell.alignment = center
        cell.border = border
        ws.cell(row=2, column=col).alignment = wrap
        ws.cell(row=2, column=col).border = border

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 26
    ws.column_dimensions["H"].width = 28
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


async def upsert_request_excel(session, request: Request, files_dir: str) -> None:
    filename = f"request_{request.id}.xlsx"
    result = await session.execute(
        select(Request)
        .where(Request.id == request.id)
        .options(
            selectinload(Request.initiator),
            selectinload(Request.department),
            selectinload(Request.cfo),
            selectinload(Request.status),
        )
        .execution_options(populate_existing=True)
    )
    req = result.scalar_one_or_none()
    if not req:
        return
    if req.description_method == "excel":
        return
    items = (
        await session.scalars(
            select(RequestItem)
            .where(RequestItem.request_id == request.id)
            .order_by(RequestItem.id)
        )
    ).all()
    attachments = (
        await session.scalars(
            select(Attachment)
            .where(Attachment.request_id == request.id)
            .order_by(Attachment.id)
        )
    ).all()
    content = build_request_xlsx(req, items, attachments)
    path = save_bytes_file(content, files_dir, filename)
    attachment = await session.scalar(
        select(Attachment).where(
            Attachment.request_id == request.id,
            Attachment.file_name == filename,
            Attachment.item_id.is_(None),
        )
    )
    if attachment:
        attachment.file_path = path
        attachment.file_type = "document"
        attachment.file_id = None
        attachment.file_unique_id = None
    else:
        session.add(
            Attachment(
                request_id=request.id,
                uploader_id=request.initiator_id,
                item_id=None,
                file_id=None,
                file_unique_id=None,
                file_name=filename,
                file_path=path,
                file_type="document",
            )
        )
